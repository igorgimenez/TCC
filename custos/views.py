
from django.shortcuts import render,redirect
from custos.forms import CustoForm,CustoEditForm 
from custos.models import  Custo
from viagem.models import  Viagem
from custos.forms import CustoFormset, CustoModelFormset
from django.http import HttpResponse
from django.db.models import Sum
from datetime import datetime

from custos.filters import CustoFilter
import xlwt

from openpyxl import Workbook
from django.contrib import messages
 
def custos(request,id):  
    custos_list = Custo.objects.filter(v_id = id)
    custo_filter = CustoFilter(request.GET, queryset=custos_list)
    v_id = id
    return render(request,"custos.html",{'filter': custo_filter, 'v_id': v_id })  
    



def criar_custo(request,id):

        viagens = Viagem.objects.filter(id=id)
     


        template_name = 'store/criar_custo.html'
        heading_message = 'Custos'

        if request.method == 'GET':
            formset = CustoFormset(request.GET or None)
        elif request.method == 'POST':
            formset = CustoFormset(request.POST)
            if formset.is_valid():
                for form in formset:
                    tipo = form.cleaned_data.get('tipo')
                    valor = form.cleaned_data.get('valor')
                    obs = form.cleaned_data.get('obs')
                    v_id =  form.cleaned_data.get('v_id')
                    if tipo:
                    
                        Custo(tipo=tipo,valor=valor,obs=obs,v_id=id).save()
                url_id= str(id)
                messages.success(request, "Custo criado com sucesso") 
                return redirect('/custos/%s'%url_id)

        return render(request, template_name, {
            'formset': formset,
            'heading': heading_message,
            'viagens':viagens,
        })


def editar(request, id):  
    custo = Custo.objects.get(id=id)  
    return render(request,'editar_custo.html', {'custo':custo})  

def update_custo(request, id):  
    custo = Custo.objects.get(id=id)  
    form = CustoEditForm(request.POST, instance = custo)
    if form.is_valid():  
        custo=form.save(commit=False)
        custo.valor = form.cleaned_data.get("valor")  
        custo.obs= form.cleaned_data.get("obs") 
        custo.updated_at = datetime.now()
        v_id = form.cleaned_data.get('v_id')
        custo.save()   
        url_id= str(v_id)
        messages.success(request, "Custo atualizado com sucesso")
        return redirect ('/custos/%s'%url_id)   
    return render(request, 'editar_custo.html', {'custo': custo})  


def excluir(request, id):  
    custo = Custo.objects.get(id=id)  
    custo.delete()  
    campo = 'v_id'
    valor = getattr(custo, campo)
    messages.success(request, "Custo excluido com sucesso")
    url_id= str(valor)
    return redirect('/custos/%s'%url_id)  



def pie_chart(request,id):
   
    v_id=id
    alimentacao=0
    transporte=0
    compras=0
    atividades=0
    passagem=0
    outros=0




    alimentacao = Custo.objects.filter(tipo="Alimentação",v_id=v_id).aggregate(TOTAL = Sum('valor'))['TOTAL']
    transporte =  Custo.objects.filter(tipo="Transporte",v_id=v_id).aggregate(TOTAL = Sum('valor'))['TOTAL']
    compras = Custo.objects.filter(tipo="Compras",v_id=v_id).aggregate(TOTAL = Sum('valor'))['TOTAL']
    atividades= Custo.objects.filter(tipo="Atividade",v_id=v_id).aggregate(TOTAL = Sum('valor'))['TOTAL']
    passagem = Custo.objects.filter(tipo="Passagem Viagem",v_id=v_id).aggregate(TOTAL = Sum('valor'))['TOTAL']
    outros = Custo.objects.filter(tipo="Outro",v_id=v_id).aggregate(TOTAL = Sum('valor'))['TOTAL']
   

    if alimentacao == None:
        alimentacao = 0
    if transporte == None:
        transporte = 0
    if compras == None:
        compras = 0
    if atividades == None:
        atividades = 0
    if passagem == None:
        passagem = 0
    if outros == None:
        outros = 0

    labels = []
    data = []
    
    
    labels.append("Alimentação")
    data.append(alimentacao)

    labels.append("Transporte")
    data.append(transporte)
 
    labels.append("Compras")
    data.append(compras)

    labels.append("Atividades")
    data.append(atividades)

    labels.append("Passagem")
    data.append(passagem)

    labels.append("Outros")
    data.append(outros)


    return render(request, 'pie_chart.html', {
        'labels': labels,
        'data': data,
    })



def exportar(request,id):

    custos_lista = Custo.objects.filter(v_id=id)
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-custos.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()
    
   
    worksheet = workbook.active
    worksheet.title = 'Custos'

   
    columns = [
        'Tipo',
        'Valor',
        'Obs',
    ]
    row_num = 1

    
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    
    for custo in custos_lista:
        row_num += 1
        
        
        row = [
    
            custo.tipo,
            custo.valor,
            custo.obs,
        ]
        
        
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response





def export_excel(request,id):

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="custos.xlsx"'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Custos')

   
    row_num = 0

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    columns = ['Tipo', 'Valor', 'Observação' ]

    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_style)

   
    font_style = xlwt.XFStyle()

    rows = Custo.objects.filter(v_id=id).values_list('tipo', 'valor', 'obs')
    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num], font_style)

    wb.save(response)
    return response